# -*- coding: utf-8 -*-
"""Excel 文件对比核心逻辑（从 ExcelComparator.py 提取，无 GUI 依赖）。"""

import datetime as dt
import logging
from enum import Enum
from pathlib import Path
from typing import Dict, List, Tuple, Union

import openpyxl as pxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, colors

logger = logging.getLogger(__name__)


def _is_na(value) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _to_compare_str(value) -> str:
    """将单元格值或列名统一转为可对比、可拼接的字符串。"""
    if _is_na(value):
        return ""
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, dt.datetime):
        if (
            value.hour == 0
            and value.minute == 0
            and value.second == 0
            and value.microsecond == 0
        ):
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


def _merge_column_labels(cols_a: List, cols_b: List) -> List[str]:
    seen = set()
    merged: List[str] = []
    for col in list(cols_a) + list(cols_b):
        label = _to_compare_str(col)
        if label not in seen:
            seen.add(label)
            merged.append(label)
    return merged


class CompareMode(str, Enum):
    DIRECT = "direct"
    KEY = "key"


def trans_none(value) -> str:
    return _to_compare_str(value)


def _parse_key_columns(keys: str) -> List[str]:
    cols = [title.strip() for title in keys.split(",") if title.strip()]
    if not cols:
        raise ValueError("主键对比模式需要指定主键列名")
    return cols


def _read_excel(path: Union[str, Path], sheet_name: str = None) -> pd.DataFrame:
    ext = Path(path).suffix.lower()
    engine = "xlrd" if ext == ".xls" else "openpyxl"
    if sheet_name:
        return pd.read_excel(path, engine=engine, sheet_name=sheet_name)
    return pd.read_excel(path, engine=engine)


def _validate_key_columns(df: pd.DataFrame, key_cols: List[str], label: str) -> None:
    missing = [col for col in key_cols if col not in df.columns]
    if missing:
        available = ", ".join(str(c) for c in df.columns)
        raise ValueError(
            "文件 {} 中不存在主键列: {}（现有列: {}）".format(label, ", ".join(missing), available)
        )


def generate_primary_key(df: pd.DataFrame, keyname: str, key_cols: List[str]) -> pd.Series:
    if len(keyname) == 0:
        keyname = "pri_key"

    df[keyname] = ""
    for title in key_cols:
        df[title] = df[title].apply(lambda x: _to_compare_str(x))
        df[keyname] += "|" + df[title]
    return df[keyname]


def get_sheet_names(path: Union[str, Path]) -> List[str]:
    """获取 Excel 文件中的所有 sheet 名称。"""
    ext = Path(path).suffix.lower()
    engine = "xlrd" if ext == ".xls" else "openpyxl"
    xls = pd.ExcelFile(path, engine=engine)
    return xls.sheet_names


def _format_diff_summary(diff_columns: List) -> str:
    if not diff_columns:
        return ""
    labels = [_to_compare_str(col) for col in diff_columns]
    if len(labels) <= 5:
        return "、".join(labels)
    return "{} 等共{}列".format("、".join(labels[:5]), len(labels))


def _classify_row_diff_type(is_only_a: bool, is_only_b: bool, has_value_diff: bool) -> str:
    if is_only_a:
        return "仅文件A"
    if is_only_b:
        return "仅文件B"
    if has_value_diff:
        return "值不同"
    return ""


def _create_diff_index_sheet(workbook: pxl.Workbook, diff_rows: List[Dict]) -> None:
    """创建差异索引 Sheet，便于在列很多时快速定位有差异的行。"""
    if not diff_rows:
        return

    index_sheet = workbook.create_sheet(title="差异索引", index=0)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    link_font = Font(color="0563C1", underline="single")

    headers = ["序号", "明细行号", "行标识", "差异类型", "差异列数", "差异列", "差异详情"]
    for col_idx, title in enumerate(headers, 1):
        cell = index_sheet.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill

    for seq, item in enumerate(diff_rows, 1):
        row_num = seq + 1
        detail_row = item["detail_row"]
        index_sheet.cell(row=row_num, column=1, value=seq)

        link_cell = index_sheet.cell(row=row_num, column=2, value=detail_row)
        link_cell.hyperlink = "#'明细数据'!A{}".format(detail_row)
        link_cell.font = link_font

        index_sheet.cell(row=row_num, column=3, value=item["row_label"])
        index_sheet.cell(row=row_num, column=4, value=item["diff_type"])
        index_sheet.cell(row=row_num, column=5, value=len(item["diff_columns"]))
        index_sheet.cell(row=row_num, column=6, value=_format_diff_summary(item["diff_columns"]))
        index_sheet.cell(row=row_num, column=7, value="; ".join(_to_compare_str(part) for part in item["diff_details"]))

    index_sheet.column_dimensions["A"].width = 8
    index_sheet.column_dimensions["B"].width = 10
    index_sheet.column_dimensions["C"].width = 24
    index_sheet.column_dimensions["D"].width = 12
    index_sheet.column_dimensions["E"].width = 10
    index_sheet.column_dimensions["F"].width = 36
    index_sheet.column_dimensions["G"].width = 60
    index_sheet.freeze_panes = "A2"


def _write_compare_result_sheet(
    result_sheet,
    data_a: List[list],
    data_b: List[list],
    all_cols: List[str],
    max_row: int,
    max_column: int,
    *,
    skip_key_in_diff_count: bool = False,
) -> Tuple[int, int, List[Dict]]:
    """写入明细对比结果，并在首列生成差异摘要。返回差异计数与差异行索引。"""
    diff_count = 0
    yellow_diff_count = 0
    diff_rows: List[Dict] = []

    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    light_green_fill = PatternFill("solid", fgColor="C6EFCE")
    light_red_fill = PatternFill("solid", fgColor="FFC7CE")
    bold_font = Font(color=colors.BLACK, bold=True)
    header_font = Font(bold=True)
    summary_header_fill = PatternFill("solid", fgColor="E2EFDA")

    summary_cell = result_sheet.cell(row=1, column=1, value="差异摘要")
    summary_cell.font = header_font
    summary_cell.fill = summary_header_fill

    data_col_offset = 2
    for col_idx, col_name in enumerate(all_cols, data_col_offset):
        cell = result_sheet.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    empty_from = 1 if skip_key_in_diff_count else 0

    for row_idx in range(max_row):
        row_a = data_a[row_idx] if row_idx < len(data_a) else []
        row_b = data_b[row_idx] if row_idx < len(data_b) else []

        row_a_all_empty = all(
            not _to_compare_str(row_a[col_idx] if col_idx < len(row_a) else "")
            for col_idx in range(empty_from, max_column)
        )
        row_b_all_empty = all(
            not _to_compare_str(row_b[col_idx] if col_idx < len(row_b) else "")
            for col_idx in range(empty_from, max_column)
        )

        is_only_a = not row_a_all_empty and row_b_all_empty
        is_only_b = row_a_all_empty and not row_b_all_empty

        row_diff_columns: List[str] = []
        row_diff_details: List[str] = []
        has_value_diff = False
        excel_row = row_idx + 2

        for col_idx in range(max_column):
            raw1 = row_a[col_idx] if col_idx < len(row_a) else ""
            raw2 = row_b[col_idx] if col_idx < len(row_b) else ""
            val1 = _to_compare_str(raw1)
            val2 = _to_compare_str(raw2)

            col_name = all_cols[col_idx] if col_idx < len(all_cols) else "列{}".format(col_idx + 1)
            cell = result_sheet.cell(row=excel_row, column=col_idx + data_col_offset)

            if skip_key_in_diff_count and col_idx == 0 and (val1 or val2):
                cell.value = val1 if val1 else val2
                if is_only_a:
                    cell.fill = light_green_fill
                    cell.font = bold_font
                elif is_only_b:
                    cell.fill = light_red_fill
                    cell.font = bold_font
            elif val1 != val2:
                diff_count += 1
                cell.value = "{}-->({})".format(val1, val2)
                row_diff_columns.append(col_name)
                row_diff_details.append("{}: {}-->{}".format(col_name, val1, val2))
                if is_only_a:
                    cell.fill = light_green_fill
                    cell.font = bold_font
                elif is_only_b:
                    cell.fill = light_red_fill
                    cell.font = bold_font
                else:
                    has_value_diff = True
                    yellow_diff_count += 1
                    cell.fill = yellow_fill
                    cell.font = bold_font
            else:
                cell.value = val1

        if row_diff_columns:
            summary_cell = result_sheet.cell(row=excel_row, column=1, value=_format_diff_summary(row_diff_columns))
            summary_cell.font = bold_font
            if is_only_a:
                summary_cell.fill = light_green_fill
            elif is_only_b:
                summary_cell.fill = light_red_fill
            else:
                summary_cell.fill = yellow_fill

            row_label = ""
            if skip_key_in_diff_count and max_column > 0:
                key_val = _to_compare_str(row_a[0]) if row_a else ""
                if not key_val and row_b:
                    key_val = _to_compare_str(row_b[0])
                row_label = key_val
            if not row_label:
                row_label = "第{}行".format(row_idx + 1)

            diff_rows.append(
                {
                    "detail_row": excel_row,
                    "row_label": row_label,
                    "diff_type": _classify_row_diff_type(is_only_a, is_only_b, has_value_diff),
                    "diff_columns": row_diff_columns,
                    "diff_details": row_diff_details,
                }
            )

    freeze_col = "C2" if skip_key_in_diff_count and max_column > 0 else "B2"
    result_sheet.freeze_panes = freeze_col
    result_sheet.column_dimensions["A"].width = 28

    return diff_count, yellow_diff_count, diff_rows


def process_file(
    file_a: Union[str, Path],
    file_b: Union[str, Path],
    keys: str,
    sheet_a: str = None,
    sheet_b: str = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """处理文件并基于主键对齐数据。

    Returns:
        Tuple[file_a_df_aligned, file_b_df_aligned, file_a_df_original, file_b_df_original]
    """
    key_cols = _parse_key_columns(keys)

    file_a_df = _read_excel(file_a, sheet_name=sheet_a)
    file_b_df = _read_excel(file_b, sheet_name=sheet_b)

    _validate_key_columns(file_a_df, key_cols, "A")
    _validate_key_columns(file_b_df, key_cols, "B")

    file_a_df_original = file_a_df.copy()
    file_b_df_original = file_b_df.copy()

    file_a_df["key1"] = generate_primary_key(df=file_a_df, keyname="key1", key_cols=key_cols)
    file_b_df["key1"] = generate_primary_key(df=file_b_df, keyname="key1", key_cols=key_cols)

    file_a_df = file_a_df.sort_values(by=key_cols).reset_index(drop=True)
    file_b_df = file_b_df.sort_values(by=key_cols).reset_index(drop=True)

    file_a_df["count"] = file_a_df.groupby(["key1"]).cumcount() + 1
    file_b_df["count"] = file_b_df.groupby(["key1"]).cumcount() + 1

    final_key_cols = ["key1", "count"]
    file_a_df["key"] = generate_primary_key(df=file_a_df, keyname="key", key_cols=final_key_cols)
    file_b_df["key"] = generate_primary_key(df=file_b_df, keyname="key", key_cols=final_key_cols)

    primary_key_list = pd.merge(file_a_df, file_b_df, on="key", how="outer", indicator=True)
    primary_key_list = primary_key_list["key"].drop_duplicates()

    file_a_df = pd.merge(primary_key_list, file_a_df, on="key", how="left")
    file_b_df = pd.merge(primary_key_list, file_b_df, on="key", how="left")

    file_a_df = file_a_df.sort_values(by=["key"]).reset_index(drop=True)
    file_b_df = file_b_df.sort_values(by=["key"]).reset_index(drop=True)

    cols_to_drop = ["key1", "count"]
    file_a_df = file_a_df.drop(columns=[c for c in cols_to_drop if c in file_a_df.columns], errors="ignore")
    file_b_df = file_b_df.drop(columns=[c for c in cols_to_drop if c in file_b_df.columns], errors="ignore")

    if "key" in file_a_df.columns:
        cols_a = ["key"] + [col for col in file_a_df.columns if col != "key"]
        file_a_df = file_a_df.reindex(columns=cols_a)

    if "key" in file_b_df.columns:
        cols_b = ["key"] + [col for col in file_b_df.columns if col != "key"]
        file_b_df = file_b_df.reindex(columns=cols_b)

    return file_a_df, file_b_df, file_a_df_original, file_b_df_original


def compare_excel_method(
    file_a: Union[str, Path],
    file_b: Union[str, Path],
    result: Union[str, Path],
    sheet_a: str = None,
    sheet_b: str = None,
) -> int:
    """逐单元格对比两个 Excel 文件，差异写入 result 文件。返回差异单元格数量。"""
    import time

    start_time = time.time()

    logger.info("开始读取 Excel 文件: file_a=%s sheet_a=%s file_b=%s sheet_b=%s", file_a, sheet_a, file_b, sheet_b)
    df_a = _read_excel(file_a, sheet_name=sheet_a)
    df_b = _read_excel(file_b, sheet_name=sheet_b)

    for col in df_a.columns:
        df_a[col] = df_a[col].apply(_to_compare_str)
    for col in df_b.columns:
        df_b[col] = df_b[col].apply(_to_compare_str)

    read_time = time.time() - start_time
    logger.info(
        "读取 Excel 完成: 耗时=%.2f秒, file_a=%dx%d, file_b=%dx%d",
        read_time,
        len(df_a),
        len(df_a.columns),
        len(df_b),
        len(df_b.columns),
    )

    max_row = max(len(df_a), len(df_b))
    max_column = max(len(df_a.columns), len(df_b.columns))

    cols_a = list(df_a.columns) if len(df_a.columns) > 0 else []
    cols_b = list(df_b.columns) if len(df_b.columns) > 0 else []
    all_cols = _merge_column_labels(cols_a, cols_b)

    data_a = df_a.values.tolist() if len(df_a) > 0 else []
    data_b = df_b.values.tolist() if len(df_b) > 0 else []

    while len(data_a) < max_row:
        data_a.append([])
    while len(data_b) < max_row:
        data_b.append([])

    result_workbook = pxl.Workbook()
    result_sheet = result_workbook.active
    result_sheet.title = "明细数据"

    compare_start = time.time()
    logger.info("开始对比数据: 总行数=%d, 总列数=%d", max_row, max_column)
    diff_count, yellow_diff_count, diff_rows = _write_compare_result_sheet(
        result_sheet,
        data_a,
        data_b,
        all_cols,
        max_row,
        max_column,
        skip_key_in_diff_count=False,
    )

    compare_time = time.time() - compare_start
    logger.info(
        "对比完成: 耗时=%.2f秒, 差异单元格=%d, 差异行=%d",
        compare_time,
        diff_count,
        len(diff_rows),
    )

    _create_diff_index_sheet(result_workbook, diff_rows)

    save_start = time.time()
    result_workbook.save(result)
    save_time = time.time() - save_start
    total_time = save_time + compare_time + read_time

    logger.info("保存结果完成: 耗时=%.2f秒, 总耗时=%.2f秒", save_time, total_time)
    return diff_count


def _create_summary_sheet(
    workbook: pxl.Workbook,
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    df_a_original: pd.DataFrame,
    df_b_original: pd.DataFrame,
    diff_count: int,
    remaining_diffs: int,
) -> None:
    """创建统计 Sheet，展示对比过程的详细差异说明。"""
    summary_sheet = workbook.create_sheet(title="对比统计")

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    total_rows = len(df_a)
    original_a_rows = len(df_a_original) if df_a_original is not None else len(df_a)
    original_b_rows = len(df_b_original) if df_b_original is not None else len(df_b)

    only_a_count = 0
    only_b_count = 0

    check_cols_a = [col for col in df_a.columns][1:] if len(df_a.columns) > 1 else []
    check_cols_b = [col for col in df_b.columns][1:] if len(df_b.columns) > 1 else []

    for idx in range(total_rows):
        if check_cols_a:
            row_a_empty = all(
                pd.isna(df_a.iloc[idx][col]) or str(df_a.iloc[idx][col]).strip() in ["", "nan", "none"]
                for col in check_cols_a
            )
        else:
            row_a_empty = True

        if check_cols_b:
            row_b_empty = all(
                pd.isna(df_b.iloc[idx][col]) or str(df_b.iloc[idx][col]).strip() in ["", "nan", "none"]
                for col in check_cols_b
            )
        else:
            row_b_empty = True

        if row_b_empty and not row_a_empty:
            only_a_count += 1
        elif row_a_empty and not row_b_empty:
            only_b_count += 1

    summary_sheet.cell(row=1, column=1, value="Excel 对比统计报告").font = title_font
    summary_sheet.merge_cells("A1:C1")

    stats = [
        ("统计项目", "数量", "说明"),
        ("一共对比数据条数", total_rows, "对齐后的总数据条数（包含补全的缺失数据）"),
        ("文件A原始数据条数", original_a_rows, "文件A原始数据条数"),
        ("文件B原始数据条数", original_b_rows, "文件B原始数据条数"),
        ("一共有差异项数", diff_count, "存在差异的单元格总数"),
        ("仅文件A数据条数", only_a_count, "仅文件A有的数据条数（文件B中不存在）- 浅绿色标记"),
        ("仅文件B数据条数", only_b_count, "仅文件B有的数据条数（文件A中不存在）- 浅红色标记"),
        ("剩余差异项数", remaining_diffs, "两个文件都有但值不同的差异项数 - 黄色标记"),
    ]

    for row_num, (item, value, desc) in enumerate(stats, start=3):
        if item == "统计项目":
            for col_num, val in enumerate([item, value, desc], start=1):
                cell = summary_sheet.cell(row=row_num, column=col_num, value=val)
                cell.font = header_font
                cell.fill = header_fill
        else:
            summary_sheet.cell(row=row_num, column=1, value=item)
            summary_sheet.cell(row=row_num, column=2, value=value)
            summary_sheet.cell(row=row_num, column=3, value=desc)

    summary_sheet.column_dimensions["A"].width = 20
    summary_sheet.column_dimensions["B"].width = 15
    summary_sheet.column_dimensions["C"].width = 50


def compare_excel_from_dataframes(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    result: Union[str, Path],
    df_a_original: pd.DataFrame = None,
    df_b_original: pd.DataFrame = None,
) -> int:
    """从已对齐的 DataFrame 对比并生成结果文件。返回差异单元格数量。"""
    import time

    start_time = time.time()

    logger.info(
        "开始从 DataFrame 对比数据: file_a=%dx%d, file_b=%dx%d",
        len(df_a),
        len(df_a.columns),
        len(df_b),
        len(df_b.columns),
    )

    for col in df_a.columns:
        df_a[col] = df_a[col].apply(_to_compare_str)
    for col in df_b.columns:
        df_b[col] = df_b[col].apply(_to_compare_str)

    max_row = max(len(df_a), len(df_b))
    max_column = max(len(df_a.columns), len(df_b.columns))

    cols_a = list(df_a.columns) if len(df_a.columns) > 0 else []
    cols_b = list(df_b.columns) if len(df_b.columns) > 0 else []
    all_cols = _merge_column_labels(cols_a, cols_b)

    data_a = df_a.values.tolist() if len(df_a) > 0 else []
    data_b = df_b.values.tolist() if len(df_b) > 0 else []

    while len(data_a) < max_row:
        data_a.append([])
    while len(data_b) < max_row:
        data_b.append([])

    result_workbook = pxl.Workbook()
    result_sheet = result_workbook.active
    result_sheet.title = "明细数据"

    compare_start = time.time()
    logger.info("开始对比数据: 总行数=%d, 总列数=%d", max_row, max_column)
    diff_count, yellow_diff_count, diff_rows = _write_compare_result_sheet(
        result_sheet,
        data_a,
        data_b,
        all_cols,
        max_row,
        max_column,
        skip_key_in_diff_count=True,
    )

    compare_time = time.time() - compare_start
    logger.info(
        "对比完成: 耗时=%.2f秒, 差异单元格=%d, 黄色差异=%d, 差异行=%d",
        compare_time,
        diff_count,
        yellow_diff_count,
        len(diff_rows),
    )

    _create_diff_index_sheet(result_workbook, diff_rows)
    _create_summary_sheet(
        result_workbook, df_a, df_b, df_a_original, df_b_original, diff_count, yellow_diff_count
    )

    save_start = time.time()
    result_workbook.save(result)
    save_time = time.time() - save_start
    total_time = save_time + compare_start - start_time

    logger.info("保存结果完成: 耗时=%.2f秒, 总耗时=%.2f秒", save_time, total_time)
    return diff_count


def compare_excel(
    file_a: Union[str, Path],
    file_b: Union[str, Path],
    result: Union[str, Path],
    mode: CompareMode = CompareMode.DIRECT,
    keys: str = "",
    sheet_a: str = None,
    sheet_b: str = None,
) -> dict:
    """
    执行 Excel 对比。

    Returns:
        {"diff_count": int, "mode": str}
    """
    file_a = Path(file_a)
    file_b = Path(file_b)
    result = Path(result)

    if mode == CompareMode.KEY:
        if not keys.strip():
            raise ValueError("主键对比模式需要指定主键列名")

        file_a_df, file_b_df, file_a_orig, file_b_orig = process_file(
            file_a, file_b, keys, sheet_a=sheet_a, sheet_b=sheet_b
        )
        diff_count = compare_excel_from_dataframes(
            file_a_df, file_b_df, result, df_a_original=file_a_orig, df_b_original=file_b_orig
        )
    else:
        diff_count = compare_excel_method(file_a, file_b, result, sheet_a=sheet_a, sheet_b=sheet_b)

    return {"diff_count": diff_count, "mode": mode.value}
